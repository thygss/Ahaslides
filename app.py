import pandas as pd
import tkinter as tk
from tkinter import filedialog
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.io as pio
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment
import datetime
import cv2
import os

# Função para selecionar o arquivo

def selecionar_arquivo():
    root = tk.Tk()
    root.withdraw() # Esconde a janela principal do tkinter
    caminho_arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    return caminho_arquivo

def limpar_dados(df, colunas_nao_remover=None):
    # Remove duplicatas
    df = df.drop_duplicates()
    
    # Se houver colunas específicas que você não quer remover, mantenha as linhas com NaN nessas colunas
    if colunas_nao_remover:
        for coluna in colunas_nao_remover:
            if coluna in df.columns: # Verifica se a coluna existe antes de tentar acessá-la
                df = df[df[coluna].notna()]
            else:
                print(f"A coluna '{coluna}' não foi encontrada no DataFrame.")
    
    return df

def analise_exploratoria(df):
    if 'Slide no' in df.columns:
        df['Slide no'] = pd.to_numeric(df['Slide no'], errors='coerce')
        
        # Agora você pode criar o histograma sem problemas
        df['Slide no'].hist()
        plt.show()
    else:
        print("A coluna 'Slide no' não foi encontrada no DataFrame.")

def criar_visualizacoes(df, pasta):
    df = df.rename(columns={'Unnamed: 0': 'Slide no', 'Unnamed: 1': 'Date created (UTC time)'})
    
    if 'Slide no' in df.columns and 'Date created (UTC time)' in df.columns:
        df['Slide no'] = pd.to_numeric(df['Slide no'], errors='coerce')
        df['Date created (UTC time)'] = pd.to_datetime(df['Date created (UTC time)'], errors='coerce')
        
        if df['Date created (UTC time)'].isnull().any():
            print("A conversão para numérico falhou para alguns valores na coluna 'Date created (UTC time).")
            return
        
        fig = px.scatter(df, x="Slide no", y="Date created (UTC time)")
        
        # Cria o caminho completo para salvar a imagem
        caminho_imagem = os.path.join(pasta, 'relatorio_visualizacao.png')
        
        # Salva a imagem do gráfico
        pio.write_image(fig, caminho_imagem)
    else:
        print("As colunas 'Slide no' e/ou 'Date created (UTC time)' não foram encontradas no DataFrame.")

def formatar_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=3, max_row=10):
        for cell in row:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.border = Border(left=Side(border_style='thin', color="000000"),
                                 right=Side(border_style='thin', color="000000"),
                                 top=Side(border_style='thin', color="000000"),
                                 bottom=Side(border_style='thin', color="000000"))
            cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(caminho_arquivo)

def salvar_em_abas(dfs_processados, caminho_arquivo):
    # Cria um novo Workbook
    wb = Workbook()
    
    # Remove a aba padrão
    wb.remove(wb.active)
    
    # Cria uma nova aba para cada DataFrame na lista
    for nome_aba, df in dfs_processados.items():
        ws = wb.create_sheet(title=nome_aba)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    
    # Salva o Workbook no arquivo especificado
    wb.save(caminho_arquivo)

def gerar_relatorio(caminho_arquivo, pasta):
    dfs = pd.read_excel(caminho_arquivo, sheet_name=None)
    dfs_processados = {nome_aba: limpar_dados(df, colunas_nao_remover=['Slide no', 'Date created (UTC time)']) for nome_aba, df in dfs.items()}
    # Aqui, em vez de chamar df_combined.to_excel, chamamos salvar_em_abas
    salvar_em_abas(dfs_processados, os.path.join(pasta, 'relatorio_melhorado.xlsx'))
    # As outras chamadas de função permanecem as mesmas
    analise_exploratoria(pd.concat(dfs_processados.values(), ignore_index=True))
    criar_visualizacoes(pd.concat(dfs_processados.values(), ignore_index=True))
    formatar_excel(os.path.join(pasta, 'relatorio_melhorado.xlsx'))
    formatar_excel(os.path.join(pasta, 'relatorio_melhorado.xlsx'))

def main():
    caminho_arquivo = selecionar_arquivo()
    if caminho_arquivo:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        pasta = 'D:\\ahaslides\\output' # Substitua pelo caminho real do diretório
        output_file = f'relatorio_melhorado_{timestamp}.xlsx'

        # Verifica se o diretório existe, caso contrário, cria
        if not os.path.exists(pasta):
            os.makedirs(pasta)
        gerar_relatorio(caminho_arquivo, pasta)
    else:
        print("Nenhum arquivo selecionado.")

if __name__ == "__main__":
    main()
    
